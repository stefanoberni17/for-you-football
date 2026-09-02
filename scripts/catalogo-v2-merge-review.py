#!/usr/bin/env python3
"""
Applica la review di Ste (xlsx compilato) alla lista candidati e produce
docs/training-catalogo-v2.json — la sorgente del catalogo v2 generato.

Regole di merge (2 set 2026):
- Ste ha corretto direttamente le colonne bianche (Gruppo, D, Liv, Attrezzatura,
  Esercizio) e scritto Note; le colonne gialle sono quasi vuote → si usa il diff.
- Gruppi nuovi scritti a mano → mappati sulle qualità canoniche (vedi GRUPPO_MAP);
  "headball" = attrezzo (il pallone Headball) → esercizio importato ma inattivo;
  "Fascia e forza" = fascia con qualità secondaria forza.
- Liv: se Ste l'ha cambiato vale il suo; altrimenti segue D (≤3 → B: "fino al 3
  lo possono fare tutti", 4 → A, 5 → PRO).
- Note "solo in coppia" → flag in_coppia; "solo con headball" → attrezzatura headball.

Uso: python3 scripts/catalogo-v2-merge-review.py <xlsx_di_ste>
"""
import json
import re
import sys

import openpyxl

ORIG = "docs/training-catalogo-v2-candidati.xlsx"
REVIEW = sys.argv[1]
BASE = json.load(open("docs/training-catalogo-v2-candidati.json", encoding="utf-8"))

GRUPPO_MAP = {
    "forza parte bassa": "forza-parte-bassa", "forza parte alta": "forza-parte-alta", "core": "core",
    "forza esplosiva": "forza-esplosiva", "pliometria estensiva": "pliometria-estensiva",
    "pliometria intensiva": "pliometria-intensiva", "velocità": "velocita", "rapidità": "velocita",
    "rapidità funzionale": "velocita", "rapidità e tecnica": "velocita", "resistenza aerobica": "resistenza-aerobica",
    "metabolico": "resistenza-metabolico", "rsa": "resistenza-rsa", "resistenza e tecnica": "resistenza-aerobica",
    "fascia": "fascia-prevenzione", "prevenzione": "fascia-prevenzione", "fascia e forza": "fascia-prevenzione",
    "tecnica palleggi": "tecnica-palleggi", "tecnica passaggi": "tecnica-passaggi", "tecnica conduzione": "tecnica-conduzione",
    "tecnica tiro": "tecnica-tiro", "tecnica visione": "tecnica-visione", "tecnica e visione": "tecnica-visione",
    "headball": "tecnica-visione", "riscaldamento": "riscaldamento", "mobilità/recupero": "mobilita-recupero",
    "test": "test", "da classificare": "da-classificare",
}
SOTTOGRUPPO = {  # etichette di Ste che aggiungono informazione oltre alla qualità canonica
    "rapidità": "rapidità", "rapidità funzionale": "rapidità funzionale", "rapidità e tecnica": "rapidità e tecnica",
    "resistenza e tecnica": "resistenza e tecnica", "fascia e forza": "fascia e forza", "prevenzione": "prevenzione",
    "headball": "headball", "tecnica e visione": "tecnica e visione",
}


def load(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Candidati"]
    hdr = [c.value for c in ws[1]]
    return {r[0].value: dict(zip(hdr, [c.value for c in r])) for r in ws.iter_rows(min_row=2) if r[0].value is not None}


orig, rev = load(ORIG), load(REVIEW)
by_num = {i + 1: row for i, row in enumerate(BASE)}
YELLOW = ("OK?", "Gruppo corretto", "Difficoltà corretta", "Livello corretto", "Azione corretta", "Note")


def liv_from_d(d):
    # Ste (2 set): la difficoltà è indipendente dal livello; fino a 3 lo possono fare tutti
    return "B" if d <= 3 else "A" if d == 4 else "PRO"


out = []
stats = {"gruppo": 0, "d": 0, "liv": 0, "attrezz": 0, "nome": 0, "note": 0, "headball": 0, "coppia": 0}
for num, base in by_num.items():
    o, n = orig[num], rev.get(num)
    row = dict(base)
    row["nome_everfit"] = base["nome"]
    row["note"] = None
    row["in_coppia"] = False
    row["attivo"] = base["azione"] == "importa"
    row["qualita_secondaria"] = None
    row["sottogruppo"] = None
    if not n:
        out.append(row)
        continue
    # nome
    if (n["Esercizio"] or "").strip() and n["Esercizio"].strip() != o["Esercizio"]:
        row["nome"] = n["Esercizio"].strip(); stats["nome"] += 1
    # gruppo: prima la colonna gialla, poi la bianca modificata
    g = (n.get("Gruppo corretto") or "").strip() or ((n["Gruppo"] or "").strip() if (n["Gruppo"] or "") != (o["Gruppo"] or "") else "")
    if g:
        key = g.lower()
        if key not in GRUPPO_MAP:
            print(f"⚠️ gruppo sconosciuto '{g}' su #{num} {row['nome']} — lasciato {row['qualita']}")
        else:
            row["qualita"] = GRUPPO_MAP[key]; stats["gruppo"] += 1
            if key in SOTTOGRUPPO:
                row["sottogruppo"] = SOTTOGRUPPO[key]
            if key == "fascia e forza":
                # Ste: spinte "overcoming isometrics" = forza parte alta E bassa (non fascia)
                row["qualita"] = "forza-parte-alta"
                row["qualita_secondaria"] = "forza-parte-bassa"
                row["sottogruppo"] = "overcoming isometrics"
            if key == "headball":
                row["attrezzatura"] = "headball"; row["attivo"] = False; stats["headball"] += 1
    # difficoltà
    d_new = n.get("Difficoltà corretta") or (n["D"] if n["D"] != o["D"] else None)
    d_changed = False
    if d_new is not None and int(d_new) != row["difficolta"]:
        row["difficolta"] = int(d_new); stats["d"] += 1; d_changed = True
    # livello
    l_new = (n.get("Livello corretto") or "").strip() or ((n["Liv"] or "").strip() if (n["Liv"] or "") != (o["Liv"] or "") else "")
    if l_new:
        row["livello_min"] = l_new.upper(); stats["liv"] += 1
    else:
        row["livello_min"] = liv_from_d(row["difficolta"])
    if row["qualita"] == "pliometria-intensiva" and row["livello_min"] == "B":
        row["livello_min"] = "A"
    # attrezzatura
    a = (n["Attrezzatura"] or "").strip()
    if a and a != (o["Attrezzatura"] or ""):
        if a.lower() == "forza funzionale":
            row["attrezzatura"] = "kettlebell"; row["sottogruppo"] = "forza funzionale"
        else:
            row["attrezzatura"] = a
        stats["attrezz"] += 1
    # azione
    az = (n.get("Azione corretta") or "").strip()
    if az:
        row["azione"] = az; row["attivo"] = az == "importa"
    # note
    note = (n.get("Note") or "").strip()
    if note and not note.startswith("ESEMPIO"):
        row["note"] = note; stats["note"] += 1
        low = note.lower()
        if "in coppia" in low:
            row["in_coppia"] = True; row["attivo"] = False; stats["coppia"] += 1  # Ste: salva ma escludi per ora
        if "solo con headball" in low or "solo con la headball" in low:
            row["attrezzatura"] = "headball"; row["attivo"] = False
        if "escludi" in low:
            row["attivo"] = False; row["azione"] = "escludi per ora"
    out.append(row)

# Esercizi base palestra (foglio CALCOLO MASSIMALI) → attrezzatura palestra anche se il nome non lo dice
BASE_PALESTRA = ("squat", "deadlift", "stacco", "hip thrust", "hip trust", "squat bulgaro", "bulgarian", "panca", "bench",
                 "shoulder press", "overhead press", "press con", "rematore", "leg press", "front squat", "sumo deadlift")
for r in out:
    n = r["nome"].lower()
    if r["qualita"] in ("forza-parte-bassa", "forza-parte-alta") and r["attrezzatura"] == "corpo libero" \
            and any(k in n for k in BASE_PALESTRA) and not any(k in n for k in ("body weight", "bodyweight", "jump", "wall", "sissy", "pistol", "iso", "atg")):
        r["attrezzatura"] = "palestra"

# Decisioni di Ste (2 set): Sumo Deadlift fuori; Squat Bulgaro = bulgarian split squat con carico (test palestra);
# Stacco Rumeno = esercizio-test della cerniera; nessuno stacco classico.
for r in out:
    n = r["nome"].lower()
    if n == "sumo deadlift":
        r["attivo"] = False; r["azione"] = "escludi (Ste: togli sumo deadlift)"
    if n == "squat bulgaro":
        r["note"] = (r.get("note") or "") + (" · " if r.get("note") else "") + "= bulgarian split squat con carico (esercizio-test batteria palestra)"
    if n == "stacco rumeno":
        r["note"] = (r.get("note") or "") + (" · " if r.get("note") else "") + "esercizio-test della cerniera (batteria palestra); nessuno stacco classico"

json.dump(out, open("docs/training-catalogo-v2.json", "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print("righe:", len(out), "| attivi:", sum(1 for r in out if r["attivo"]), "| stats:", stats)
from collections import Counter
print("qualità:", Counter(r["qualita"] for r in out).most_common())
