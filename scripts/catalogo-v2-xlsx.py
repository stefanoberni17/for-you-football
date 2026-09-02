#!/usr/bin/env python3
"""
Genera docs/training-catalogo-v2-candidati.xlsx dalla lista candidati JSON:
un foglio da rivedere riga per riga, con colonne gialle da compilare
(OK?, gruppo/difficoltà/livello corretti, note) e menu a tendina.

Uso: python3 scripts/catalogo-v2-xlsx.py
"""
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

rows = json.load(open("docs/training-catalogo-v2-candidati.json", encoding="utf-8"))

LABEL = {
    "forza-parte-bassa": "Forza parte bassa", "forza-parte-alta": "Forza parte alta", "core": "Core",
    "forza-esplosiva": "Forza esplosiva", "pliometria-estensiva": "Pliometria estensiva",
    "pliometria-intensiva": "Pliometria intensiva", "velocita": "Velocità", "resistenza-aerobica": "Resistenza aerobica",
    "resistenza-metabolico": "Metabolico", "resistenza-rsa": "RSA", "fascia-prevenzione": "Fascia",
    "tecnica-palleggi": "Tecnica palleggi", "tecnica-passaggi": "Tecnica passaggi", "tecnica-conduzione": "Tecnica conduzione",
    "tecnica-tiro": "Tecnica tiro", "tecnica-visione": "Tecnica visione", "riscaldamento": "Riscaldamento",
    "mobilita-recupero": "Mobilità/recupero", "test": "Test", "da-classificare": "Da classificare",
}
GRUPPI = list(LABEL.values())

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3D2B")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
EDIT_FILL = PatternFill("solid", fgColor="FFF2A8")
GROUP_FILL = PatternFill("solid", fgColor="E8F0EA")
BODY = Font(name=FONT, size=10)
LINK = Font(name=FONT, size=10, color="0563C1", underline="single")

wb = Workbook()

# ── Foglio Legenda ───────────────────────────────────────────────────────────
lg = wb.active
lg.title = "Legenda"
lines = [
    ("Catalogo v2 — lista candidati per la review", True),
    ("", False),
    ("Foglio 'Candidati': una riga per esercizio, ordinata per gruppo (qualità) e difficoltà.", False),
    ("Le colonne BIANCHE sono la proposta automatica (da tag e nome): non serve toccarle.", False),
    ("Le colonne GIALLE sono da compilare — solo dove vuoi cambiare qualcosa:", False),
    ("   • OK? — 'sì' se la riga va bene così, 'no' se va corretta o esclusa", False),
    ("   • Gruppo corretto / Difficoltà corretta / Livello corretto — menu a tendina", False),
    ("   • Azione corretta — importa / escludi / unisci a… (indica con chi nelle note)", False),
    ("   • Note — qualsiasi cosa da considerare: varianti, esecuzione, quando usarlo, dubbi", False),
    ("", False),
    ("Riga di esempio compilata (prima riga del foglio Candidati):", False),
    ("   OK? = no · Gruppo corretto = Forza esplosiva · Difficoltà corretta = 3 · Note = 'solo con tecnica di atterraggio pulita'", False),
    ("", False),
    ("Colonne: Gruppo = qualità v2 · Attrezzatura · D = difficoltà 1-5 · Liv = livello minimo B/A/PRO · Unità · Per lato · Video (link) · Fonte · Azione proposta", False),
    ("Se una riga è complessa da spiegare qui, scrivi 'vedi chat' nelle note e ne parliamo.", False),
]
for i, (t, b) in enumerate(lines, 1):
    c = lg.cell(i, 1, t)
    c.font = Font(name=FONT, size=12 if b else 10, bold=b)
lg.column_dimensions["A"].width = 120

# ── Foglio Candidati ─────────────────────────────────────────────────────────
ws = wb.create_sheet("Candidati")
headers = ["#", "Gruppo", "Esercizio", "Attrezzatura", "D", "Liv", "Unità", "Per lato", "Video", "Fonte", "Azione proposta",
           "OK?", "Gruppo corretto", "Difficoltà corretta", "Livello corretto", "Azione corretta", "Note"]
EDIT_COLS = {12, 13, 14, 15, 16, 17}
for j, h in enumerate(headers, 1):
    c = ws.cell(1, j, h)
    c.font = HEAD_FONT
    c.fill = HEAD_FILL
    c.alignment = Alignment(vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 30

prev_group = None
for i, r in enumerate(rows, 2):
    g = LABEL[r["qualita"]]
    vals = [i - 1, g, r["nome"], r["attrezzatura"], r["difficolta"], r["livello_min"], r["unita"],
            "sì" if r["per_lato"] else "", r["video"] or "", r["fonte"], r["azione"], "", "", "", "", "", ""]
    for j, v in enumerate(vals, 1):
        c = ws.cell(i, j, v)
        c.font = BODY
        c.alignment = Alignment(vertical="center", wrap_text=(j in (3, 17)))
        if j in EDIT_COLS:
            c.fill = EDIT_FILL
        elif g != prev_group and j <= 11:
            c.fill = GROUP_FILL
    if r["video"]:
        c = ws.cell(i, 9)
        c.value = "video"
        c.hyperlink = r["video"]
        c.font = LINK
    prev_group = g

# Riga di esempio compilata (prima riga)
ws.cell(2, 12, "no")
ws.cell(2, 13, "Forza esplosiva")
ws.cell(2, 14, 3)
ws.cell(2, 17, "ESEMPIO — solo con tecnica di atterraggio pulita (cancella o sovrascrivi)")
for j in (12, 13, 14, 17):
    ws.cell(2, j).font = Font(name=FONT, size=10, italic=True, color="7A6A00")

last = len(rows) + 1
dv_ok = DataValidation(type="list", formula1='"sì,no"', allow_blank=True)
dv_gr = DataValidation(type="list", formula1='"' + ",".join(GRUPPI) + '"', allow_blank=True)
dv_d = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
dv_l = DataValidation(type="list", formula1='"B,A,PRO"', allow_blank=True)
dv_a = DataValidation(type="list", formula1='"importa,escludi,unisci a…,già in v1"', allow_blank=True)
for dv, col in ((dv_ok, "L"), (dv_gr, "M"), (dv_d, "N"), (dv_l, "O"), (dv_a, "P")):
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{last}")

widths = [5, 20, 46, 14, 4, 5, 9, 8, 7, 16, 18, 6, 20, 10, 9, 14, 50]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = "D2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last}"

out = "docs/training-catalogo-v2-candidati.xlsx"
wb.save(out)
print(f"{out}: {len(rows)} righe")
